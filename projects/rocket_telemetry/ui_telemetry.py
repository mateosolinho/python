import os
import cv2
import pytesseract
import re
import numpy as np
from concurrent.futures import ThreadPoolExecutor
from collections import deque
import csv
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import tkinter.ttk as ttk 

class TelemetryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Rocket Telemetry")
        self.root.geometry("1540x800+0+0")
        
        # Variables
        self.video_path = ""
        self.canvas = None
        self.frame_label = None
        self.cap = None  # Para manejar la captura del video
        self.frame_counter = 0
        self.speed_buffer = deque(maxlen=5)
        self.altitude_buffer = deque(maxlen=5)

        # Crear un archivo .xlsx para almacenar los datos
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(['Time (seconds)', 'Speed', 'Altitude'])

        # ===================================================================================================

        # Definir el layout usando grid
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=0, column=0, columnspan=2, pady=10)
        
        # Botones alineados horizontalmente
        tk.Button(button_frame, text="Cargar Video", command=self.load_video).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Iniciar", command=self.run_telemetry).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Reset", command=self.reset_all).pack(side=tk.LEFT, padx=10)
        
        # ===================================================================================================
        
        # Fila 1 - Frame para visualizar el vídeo
        video_frame = tk.Frame(self.root, width=1280, height=720, bg="black")
        video_frame.grid(row=1, column=0, padx=10, pady=10)
        
        # Canvas para mostrar el vídeo
        self.canvas = tk.Canvas(video_frame, width=1280, height=720, bg="black")
        self.canvas.pack()
        
        # ===================================================================================================
        
        telemetry_frame = tk.Frame(self.root)
        telemetry_frame.grid(row=1, column=1, padx=10, pady=10)

        # Establecer un ancho fijo para las etiquetas
        label_width = 20  # Ajusta este valor según sea necesario

        # Etiquetas para mostrar la telemetría (Velocidad, Altitud, Tiempo)
        self.speed_label = tk.Label(telemetry_frame, text="Velocidad: ---", font=("Helvetica", 16), width=label_width, anchor='w')
        self.speed_label.grid(row=0, column=0, pady=10)

        self.altitude_label = tk.Label(telemetry_frame, text="Altitud: ---", font=("Helvetica", 16), width=label_width, anchor='w')
        self.altitude_label.grid(row=1, column=0, pady=10)

        self.time_label = tk.Label(telemetry_frame, text="Tiempo: ---", font=("Helvetica", 16), width=label_width, anchor='w')
        self.time_label.grid(row=2, column=0, pady=10)
        
        # ===================================================================================================

        # Etiqueta para el tiempo inicial
        initial_time_label = tk.Label(button_frame, text="T - Inicial:")
        initial_time_label.pack(side=tk.LEFT, padx=10)

        # Variables para los desplegables del tiempo inicial
        self.initial_hours = tk.StringVar(value="00")
        self.initial_minutes = tk.StringVar(value="00")
        self.initial_seconds = tk.StringVar(value="00")

        # Desplegables para horas, minutos y segundos del tiempo inicial usando Combobox
        initial_hour_options = [f"{i:02d}" for i in range(24)]
        initial_minute_options = [f"{i:02d}" for i in range(60)]
        initial_second_options = [f"{i:02d}" for i in range(60)]

        initial_hour_menu = ttk.Combobox(button_frame, textvariable=self.initial_hours, values=initial_hour_options, state='readonly', width=3)
        initial_minute_menu = ttk.Combobox(button_frame, textvariable=self.initial_minutes, values=initial_minute_options, state='readonly', width=3)
        initial_second_menu = ttk.Combobox(button_frame, textvariable=self.initial_seconds, values=initial_second_options, state='readonly', width=3)

        # Establecer valores predeterminados
        initial_hour_menu.current(0)  # 00
        initial_minute_menu.current(0)  # 00
        initial_second_menu.current(0)  # 00

        # Empaquetar los Combobox del tiempo inicial
        initial_hour_menu.pack(side=tk.LEFT)
        initial_minute_menu.pack(side=tk.LEFT)
        initial_second_menu.pack(side=tk.LEFT)
        
        # ===================================================================================================
        
        time_label = tk.Label(button_frame, text="T + Objetivo:")
        time_label.pack(side=tk.LEFT, padx=10)

        # Variables para los desplegables
        self.hours = tk.StringVar(value="00")
        self.minutes = tk.StringVar(value="00")
        self.seconds = tk.StringVar(value="00")

        # Desplegables para horas, minutos y segundos usando Combobox
        hour_options = [f"{i:02d}" for i in range(24)]
        minute_options = [f"{i:02d}" for i in range(60)]
        second_options = [f"{i:02d}" for i in range(60)]

        hour_menu = ttk.Combobox(button_frame, textvariable=self.hours, values=hour_options, state='readonly', width=3)
        minute_menu = ttk.Combobox(button_frame, textvariable=self.minutes, values=minute_options, state='readonly', width=3)
        second_menu = ttk.Combobox(button_frame, textvariable=self.seconds, values=second_options, state='readonly', width=3)

        # Establecer valores predeterminados
        hour_menu.current(0)  # 00
        minute_menu.current(0)  # 00
        second_menu.current(0)  # 00

        hour_menu.pack(side=tk.LEFT)
        minute_menu.pack(side=tk.LEFT)
        second_menu.pack(side=tk.LEFT)
        
        # ===================================================================================================
    
    def get_video_info(self, video_path):
        file_name = os.path.basename(video_path)  # Nombre del archivo
        file_size = os.path.getsize(video_path)  # Tamaño en bytes

        # Usar OpenCV para obtener la duración
        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS)  # Frames por segundo
        frame_count = cap.get(cv2.CAP_PROP_FRAME_COUNT)  # Número total de frames
        duration = frame_count / fps 
        print(file_name)
        print(file_size)
        print(duration)
        
    def reset_all(self):
        # Reiniciar los valores de los combobox a "00"
        self.initial_hours.set("00")
        self.initial_minutes.set("00")
        self.initial_seconds.set("00")
        self.hours.set("00")
        self.minutes.set("00")
        self.seconds.set("00")
        cv2.destroyAllWindows()
        
    def load_video(self):
        self.video_path = filedialog.askopenfilename(title="Seleccionar video", filetypes=[("MP4 files", "*.mp4")])
        if not self.video_path:
            messagebox.showerror("Error", "No se seleccionó ningún archivo.")
        else:
            self.get_video_info(self.video_path)
            messagebox.showinfo("Video Cargado", f"Video cargado correctamente: {self.video_path}")

    def cleanup(self, im):
        arr = np.array(im, dtype=float)
        if arr.mean(axis=-1).max() < 200:
            arr[:] = 0
        else:
            arr -= np.median(arr) + 5
            arrmax = arr.max(axis=(0, 1))
            if all(arrmax != 0):
                arr *= 255 / arrmax
            arr = arr.clip(0, 255)
        return arr.astype(np.uint8)

    def extract_text(self, image_region):
        custom_config = r'--oem 3 --psm 6'
        return pytesseract.image_to_string(image_region, config=custom_config)


    def convert_to_milliseconds(self):
        # Convierte HH:MM:SS a milisegundos
        hours = int(self.initial_hours.get())
        minutes = int(self.initial_minutes.get())
        seconds = int(self.initial_seconds.get())
        total_milliseconds = (hours * 3600 + minutes * 60 + seconds) * 1000
        return total_milliseconds

    def start_process(self):
        # Verifica que la ruta del video no esté vacía
        if not self.video_path:  # Asumiendo que self.video_path es donde se guarda la ruta del video
            messagebox.showerror("Error", "Por favor, selecciona un video.")
            return

        # Convertir el tiempo inicial a milisegundos
        self.initial_time_ms = self.convert_to_milliseconds()

    def run_telemetry(self):
        if not self.video_path:
            messagebox.showerror("Error", "Cargue un video antes de iniciar.")
            return

        self.start_process()

        # Obtener el tiempo objetivo del desplegable
        self.target_time = f"{self.hours.get()}:{self.minutes.get()}:{self.seconds.get()}"
        
        # Validar el formato del tiempo
        if not re.match(r'^\d{2}:\d{2}:\d{2}$', self.target_time):
            messagebox.showerror("Error", "Por favor, ingrese un tiempo válido en el formato HH:MM:SS.")
            return

        self.cap = cv2.VideoCapture(self.video_path)
        self.cap.set(cv2.CAP_PROP_POS_MSEC, self.initial_time_ms)
        self.frame_counter = 0
        self.time_reached = False
        self.process_frame()

    def save_to_excel(self, speed, altitude, time):
        # Verifica si el archivo ya existe
        file_exists = False
        try:
            with open("telemetry_data.xlsx", "r") as f:
                file_exists = True
        except FileNotFoundError:
            pass

        if file_exists:
            # Abre el archivo existente y agrega los nuevos datos
            workbook = load_workbook("telemetry_data.xlsx")
            sheet = workbook.active
        else:
            # Crea un nuevo archivo
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Time (seconds)', 'Speed', 'Altitude'])

        # Agrega los nuevos datos
        sheet.append([time, speed, altitude])
        
        # Guarda el archivo
        workbook.save("telemetry_data.xlsx")

    def process_frame(self):
        if self.cap.isOpened():
            ret, frame = self.cap.read()
            if not ret:
                return
            
            # Solo procesar cada 5 fotogramas
            if self.frame_counter % 5 == 0:
                rect_start_x = int(0.8 * 1750)
                rect_start_y = int(0.8 * 1135)
                rect_width = 230
                rect_height = 35
                altitude_rect_start_y = rect_start_y + rect_height
                time_rect_start_x = 905
                time_rect_start_y = 950
                time_rect_width = 155
                time_rect_height = 45

                # Extrae las regiones
                speed_region = frame[rect_start_y:rect_start_y + rect_height, rect_start_x:rect_start_x + rect_width]
                altitude_region = frame[altitude_rect_start_y:altitude_rect_start_y + rect_height, rect_start_x:rect_start_x + rect_width]
                time_region = frame[time_rect_start_y:time_rect_start_y + time_rect_height, time_rect_start_x:time_rect_start_x + time_rect_width]

                # Limpia las regiones
                cleaned_speed_region = self.cleanup(speed_region)
                cleaned_altitude_region = self.cleanup(altitude_region)
                cleaned_time_region = self.cleanup(time_region)

                # Extracción de texto en paralelo
                with ThreadPoolExecutor() as executor:
                    speed_future = executor.submit(self.extract_text, cleaned_speed_region)
                    altitude_future = executor.submit(self.extract_text, cleaned_altitude_region)
                    time_future = executor.submit(self.extract_text, cleaned_time_region)

                    speed_text = speed_future.result()
                    altitude_text = altitude_future.result()
                    time_text = time_future.result()

                # Procesar los textos
                speed_numbers = re.findall(r'\d+', speed_text.replace('"', '').replace("'", '').strip())
                altitude_numbers = re.findall(r'\d+', altitude_text.replace('"', '').replace("'", '').strip())
                time_text = re.sub(r'[^0-9:]', '', time_text)

                if speed_numbers:
                    speed_value = int(speed_numbers[0])
                    self.speed_buffer.append(speed_value)
                else:
                    speed_value = "No speed detected"

                if altitude_numbers:
                    altitude_value = int(altitude_numbers[0])
                    self.altitude_buffer.append(altitude_value)
                else:
                    altitude_value = "No altitude detected"

                self.avg_altitude = altitude_value if isinstance(altitude_value, int) else altitude_value
                self.avg_speed = speed_value if isinstance(speed_value, int) else speed_value
                
                self.speed_label.config(text=f"Velocidad: {self.avg_speed}")
                self.altitude_label.config(text=f"Altitud: {self.avg_altitude}")
                self.time_label.config(text=f"Tiempo: {time_text}")
                
                self.save_to_excel(self.avg_speed, self.avg_altitude, time_text)
                    
                if time_text == self.target_time:
                    self.time_reached = True
                    messagebox.showinfo("Tiempo Alcanzado", f"Se alcanzó el tiempo {time_text}. Datos guardados en el archivo 'telemetry_data.xlsx'.")
                    self.cap.release()  # Libera el video
                    return

                # Dibujar los rectángulos y valores en el frame
                cv2.rectangle(frame, (rect_start_x, rect_start_y), (rect_start_x + rect_width, rect_start_y + rect_height), (0, 255, 0), 2)
                cv2.rectangle(frame, (rect_start_x, altitude_rect_start_y), (rect_start_x + rect_width, altitude_rect_start_y + rect_height), (255, 0, 0), 2)
                cv2.putText(frame, f"Speed: {self.avg_speed}", (rect_start_x, rect_start_y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
                cv2.putText(frame, f"Altitude: {self.avg_altitude}", (rect_start_x, altitude_rect_start_y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 2)

            # Ajustar el tamaño del frame al tamaño del canvas
            frame_resized = cv2.resize(frame, (self.canvas.winfo_width(), self.canvas.winfo_height()))

            # Convertir el frame a imagen de PIL para mostrar en el Canvas de Tkinter
            frame_rgb = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)
            frame_pil = Image.fromarray(frame_rgb)
            self.frame_tk = ImageTk.PhotoImage(image=frame_pil)  # Mantener la referencia
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.frame_tk)
            self.root.update()

            self.frame_counter += 1
            self.root.after(1, self.process_frame)  # Procesa el siguiente fotograma tras un pequeño retraso


# Uso del programa
if __name__ == "__main__":
    root = tk.Tk()
    app = TelemetryApp(root)
    root.mainloop()
